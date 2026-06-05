# handlers/declarant.py
import os
import asyncio
from datetime import datetime
import openpyxl
from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.types import Message, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from num2words import num2words
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
from aiogram.filters.callback_data import CallbackData
from database.models import Company, Contract
from services.document_gen import generate_st1_file, generate_3qadam_file, generate_cmr_file, generate_spec_file

from database.models import Invoice, InvoiceItem

declarant_router = Router()

class InvoiceListCb(CallbackData, prefix="inv_list"):
    invoice_id: int

class CompanyCb(CallbackData, prefix="comp"):
    action: str  # 'sender' или 'consignee'
    invoice_id: int
    company_id: int

class ContractCb(CallbackData, prefix="cont"):
    invoice_id: int
    contract_id: int

class DocCb(CallbackData, prefix="doc"):
    doc_type: str # 'cmr', 'st1', '3qadam', 'spec'
    invoice_id: int

class ContractListCb(CallbackData, prefix="cont_list"):
    contract_id: int

class TemplateCb(CallbackData, prefix="tpl"):
    action: str

# Расширяем FSM для полного заполнения
class InvoiceProcess(StatesGroup):
    waiting_for_number = State()
    waiting_for_date = State()
    waiting_for_transport_number = State()  # Шаг 1: Номер машины
    waiting_for_transport_driver = State()  # Шаг 2: Водитель
    waiting_for_tir = State()
    waiting_for_destination = State()
    waiting_for_excel = State()


# 1. Нажатие на кнопку в меню -> ВЫВОД СПИСКА КОНТРАКТОВ
@declarant_router.message(Command("invoices"))
@declarant_router.message(F.text == "📋 Список инвойсов")
async def show_contracts_for_invoices(message: Message, state: FSMContext, session, user_role: str):
    if user_role not in ['admin', 'declarant']: return
    await state.clear()

    # Достаем последние 15 контрактов
    stmt = select(Contract).order_by(Contract.id.desc()).limit(15)
    contracts = (await session.scalars(stmt)).all()

    if not contracts:
        await message.answer("📭 В базе пока нет добавленных контрактов.")
        return

    # Собираем контракты в Inline-кнопки
    builder = []
    for cont in contracts:
        builder.append([
            InlineKeyboardButton(
                text=f"📜 Контракт № {cont.contract_number}",
                callback_data=ContractListCb(contract_id=cont.id).pack()
            )
        ])

    kb = InlineKeyboardMarkup(inline_keyboard=builder)

    await message.answer(
        "📋 **Архив документов:**\n"
        "Сначала выберите **Контракт**, чтобы посмотреть привязанные к нему инвойсы:",
        reply_markup=kb
    )


# 2. Нажатие на Контракт -> ВЫВОД СПИСКА ИНВОЙСОВ
@declarant_router.callback_query(ContractListCb.filter())
async def show_invoices_by_contract(call: CallbackQuery, callback_data: ContractListCb, session):
    contract_id = callback_data.contract_id

    # Ищем инвойсы ТОЛЬКО для выбранного контракта
    stmt = select(Invoice).where(
        Invoice.contract_id == contract_id,
        Invoice.status == 'completed'
    ).order_by(Invoice.id.desc()).limit(15)

    invoices = (await session.scalars(stmt)).all()

    if not invoices:
        # Если инвойсов нет, показываем всплывающее уведомление, не меняя меню
        await call.answer("К этому контракту еще не привязаны готовые инвойсы.", show_alert=True)
        return

    # Собираем инвойсы
    builder = []
    for inv in invoices:
        date_str = inv.invoice_date.strftime('%d.%m.%Y') if inv.invoice_date else 'Без даты'
        builder.append([
            InlineKeyboardButton(
                text=f"📄 № {inv.invoice_number} (от {date_str})",
                callback_data=InvoiceListCb(invoice_id=inv.id).pack()
            )
        ])

    kb = InlineKeyboardMarkup(inline_keyboard=builder)

    # Достаем сам контракт, чтобы красиво написать заголовок
    contract = await session.get(Contract, contract_id)

    await call.message.edit_text(
        f"📂 **Инвойсы по контракту № {contract.contract_number}:**\n"
        "Выберите нужный инвойс для скачивания документов:",
        reply_markup=kb
    )
    await call.answer()


# 1. Запуск
@declarant_router.message(Command("new_invoice"))
@declarant_router.message(F.text == "📄 Новый инвойс")
async def start_new_invoice(message: Message, state: FSMContext, user_role: str):
    if user_role not in ['admin', 'declarant']: return
    await state.clear()  # Сброс старых состояний
    await message.answer("📄 Введите **номер** нового инвойса:")
    await state.set_state(InvoiceProcess.waiting_for_number)


# 2. Ловим номер -> Спрашиваем дату
@declarant_router.message(InvoiceProcess.waiting_for_number)
async def process_invoice_number(message: Message, state: FSMContext, session, user_db_id: int):
    invoice_number = message.text.strip()

    new_invoice = Invoice(
        invoice_number=invoice_number,
        created_by=user_db_id,
        status='draft'
    )
    session.add(new_invoice)
    await session.commit()

    await state.update_data(invoice_id=new_invoice.id)

    await message.answer(
        f"✅ Инвойс {invoice_number} начат.\n\n"
        f"📅 Введите **дату инвойса** в формате ДД.ММ.ГГГГ (например: `01.06.2026`):"
    )
    await state.set_state(InvoiceProcess.waiting_for_date)


# 3. Ловим дату -> Спрашиваем НОМЕР ТРАНСПОРТА (ИСПРАВЛЕНО)
@declarant_router.message(InvoiceProcess.waiting_for_date)
async def process_invoice_date(message: Message, state: FSMContext, session):
    date_str = message.text.strip()

    try:
        try:
            parsed_date = datetime.strptime(date_str, "%d.%m.%Y").date()
        except ValueError:
            parsed_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        await message.answer("⚠️ Неверный формат. Введите в формате ДД.ММ.ГГГГ:")
        return

    user_data = await state.get_data()
    invoice = await session.get(Invoice, user_data['invoice_id'])
    invoice.invoice_date = parsed_date
    await session.commit()

    await message.answer("🚛 Введите **регистрационный номер транспорта** (Тягач / Прицеп):")
    await state.set_state(InvoiceProcess.waiting_for_transport_number)


# 4. Ловим номер транспорта -> Спрашиваем ВОДИТЕЛЯ (НОВЫЙ ШАГ)
@declarant_router.message(InvoiceProcess.waiting_for_transport_number)
async def process_transport_number(message: Message, state: FSMContext, session):
    transport_number = message.text.strip()

    user_data = await state.get_data()
    invoice = await session.get(Invoice, user_data['invoice_id'])
    invoice.transport_number = transport_number  # Сохраняем в правильное поле
    await session.commit()

    await message.answer("👤 Введите **ФИО водителя** (и данные паспорта, если необходимо для CMR):")
    await state.set_state(InvoiceProcess.waiting_for_transport_driver)


# 4.5. Ловим водителя -> Спрашиваем TIR
@declarant_router.message(InvoiceProcess.waiting_for_transport_driver)
async def process_transport_driver(message: Message, state: FSMContext, session):
    transport_driver = message.text.strip()

    user_data = await state.get_data()
    invoice = await session.get(Invoice, user_data['invoice_id'])
    invoice.transport_driver = transport_driver  # Сохраняем в правильное поле
    await session.commit()

    await message.answer("📝 Введите номер книжки **TIR Carnet** (или напишите 'нет'):")
    await state.set_state(InvoiceProcess.waiting_for_tir)


# 5. Ловим TIR -> Спрашиваем место выгрузки
@declarant_router.message(InvoiceProcess.waiting_for_tir)
async def process_tir(message: Message, state: FSMContext, session):
    tir_data = message.text.strip()

    user_data = await state.get_data()
    invoice = await session.get(Invoice, user_data['invoice_id'])
    invoice.transport_tir_carnet = tir_data if tir_data.lower() != 'нет' else None
    await session.commit()

    await message.answer("📍 Введите **место назначения / выгрузки** (например: `г. Ташкент, Таможенный пост...`):")
    await state.set_state(InvoiceProcess.waiting_for_destination)


# 6. Ловим место выгрузки -> Спрашиваем Excel
# 6. Ловим место выгрузки -> Спрашиваем Excel и даем кнопку шаблона
@declarant_router.message(InvoiceProcess.waiting_for_destination)
async def process_destination(message: Message, state: FSMContext, session):
    destination = message.text.strip()

    user_data = await state.get_data()
    invoice = await session.get(Invoice, user_data['invoice_id'])
    invoice.final_destination = destination
    await session.commit()

    # Создаем кнопку для скачивания шаблона
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Скачать шаблон Excel", callback_data=TemplateCb(action="download").pack())]
    ])

    await message.answer(
        "📎 Отлично! Логистическая информация сохранена.\n\n"
        "Теперь отправьте **Excel-файл** с перечнем товаров (.xlsx).\n\n"
        "Если у вас нет под рукой пустого бланка, скачайте его по кнопке ниже:",
        reply_markup=kb
    )
    await state.set_state(InvoiceProcess.waiting_for_excel)

# Функция для парсинга Excel
def parse_excel_and_calculate(filepath, invoice_id):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    items_to_create = []
    grand_total = 0.0

    count = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        qty = float(row[2]) if row[2] else 0.0
        price = float(row[4]) if row[4] else 0.0
        total = round(qty * price, 2)
        grand_total += total

        item = InvoiceItem(
            invoice_id=invoice_id,
            item_number=count,
            item_name=row[0],
            hs_code=str(row[1]) if row[1] else '',
            quantity=qty,
            unit=row[3] if row[3] else 'шт',
            price=price,
            total=total,
            package_quantity=int(row[5]) if row[5] else 0,
            package_type=str(row[6]) if row[6] else '',
            package_code=str(row[7]) if row[7] else '',
            piece_quantity=str(row[8]) if row[8] else 0,
            net_weight=float(row[9]) if row[9] else 0.0,
            gross_weight=float(row[10]) if row[10] else 0.0
        )
        items_to_create.append(item)
        count += 1
    return items_to_create, grand_total


# 7. Ловим файл и сохраняем товары
@declarant_router.message(InvoiceProcess.waiting_for_excel, F.document)
async def process_excel_file(message: Message, bot: Bot, state: FSMContext, session):
    document = message.document

    if not document.file_name.endswith('.xlsx'):
        await message.answer("⚠️ Пожалуйста, отправьте файл в формате .xlsx")
        return

    user_data = await state.get_data()
    invoice_id = user_data['invoice_id']

    temp_dir = "temp_files"
    os.makedirs(temp_dir, exist_ok=True)
    file_path = os.path.join(temp_dir, f"{document.file_id}.xlsx")

    processing_msg = await message.answer("⏳ Скачиваю и обрабатываю файл...")

    try:
        file = await bot.get_file(document.file_id)
        await bot.download_file(file.file_path, destination=file_path)

        # Читаем товары и считаем сумму
        items_to_create, grand_total = await asyncio.to_thread(
            parse_excel_and_calculate, file_path, invoice_id
        )

        session.add_all(items_to_create)

        # Обновляем инвойс
        stmt = select(Invoice).where(Invoice.id == invoice_id)
        invoice = await session.scalar(stmt)
        invoice.grand_total = grand_total

        # --- НОВАЯ ЛОГИКА: СУММА ПРОПИСЬЮ ---
        try:
            # По умолчанию ставим USD (доллары).
            # Позже, когда добавим контракты, сможем брать валюту оттуда.
            words = num2words(grand_total, lang='ru', to='currency', currency='USD')
            invoice.grand_total_words = words.capitalize()
        except Exception as e:
            print(f"Ошибка num2words: {e}")
            invoice.grand_total_words = ""
        # ------------------------------------

        await session.commit()
        os.remove(file_path)

        # --- ВМЕСТО ПРОСТОГО СООБЩЕНИЯ ГЕНЕРИРУЕМ КНОПКИ ---

        # Ищем все компании пользователя (и админа) для выбора Отправителя
        # Ищем только ОТПРАВИТЕЛЕЙ
        companies_stmt = select(Company).where(
            Company.company_type == 'sender'  # <--- Фильтр
        )
        companies = (await session.scalars(companies_stmt)).all()

        if not companies:
            await processing_msg.edit_text(
                "✅ Товары загружены, но в базе нет Отправителей!\n"
                "Добавьте их командой /add_company, а затем начните заново."
            )
            return

        builder = []
        for comp in companies:
            builder.append([
                InlineKeyboardButton(
                    text=f"📤 {comp.name}",
                    callback_data=CompanyCb(action="sender", invoice_id=invoice_id, company_id=comp.id).pack()
                )
            ])
        kb = InlineKeyboardMarkup(inline_keyboard=builder)

        await processing_msg.edit_text(
            f"✅ Файл успешно обработан! Общая сумма: {grand_total} USD.\n\n"
            f"👇 **Шаг 1/3: Выберите Отправителя (Sender)**",
            reply_markup=kb
        )

    except Exception as e:
        await session.rollback()
        await processing_msg.edit_text(f"❌ Произошла ошибка при обработке файла: {e}")
        if os.path.exists(file_path):
            os.remove(file_path)
    finally:
        await state.clear()


# --- ОБРАБОТКА НАЖАТИЙ INLINE КНОПОК ---

# 1. Выбрали Отправителя -> Спрашиваем Получателя
@declarant_router.callback_query(CompanyCb.filter(F.action == "sender"))
async def select_sender(call: CallbackQuery, callback_data: CompanyCb, session):
    invoice = await session.get(Invoice, callback_data.invoice_id)
    invoice.company_id = callback_data.company_id
    await session.commit()

    # Ищем только ПОЛУЧАТЕЛЕЙ
    companies_stmt = select(Company).where(Company.company_type == 'consignee')  # <--- Фильтр
    companies = (await session.scalars(companies_stmt)).all()

    if not companies:
        await call.message.edit_text("⚠️ В базе нет Получателей! Добавьте через /add_company")
        return

    builder = []
    for comp in companies:
        builder.append([
            InlineKeyboardButton(
                text=f"📥 {comp.name}",
                callback_data=CompanyCb(action="consignee", invoice_id=callback_data.invoice_id,
                                        company_id=comp.id).pack()
            )
        ])
    kb = InlineKeyboardMarkup(inline_keyboard=builder)

    await call.message.edit_text(
        "👇 **Шаг 2/3: Выберите Получателя (Consignee)**",
        reply_markup=kb
    )
    await call.answer()


# 2. Выбрали Получателя -> Спрашиваем Контракт
# 2. Выбрали Получателя -> Спрашиваем Контракт
@declarant_router.callback_query(CompanyCb.filter(F.action == "consignee"))
async def select_consignee(call: CallbackQuery, callback_data: CompanyCb, session):
    # Сохраняем получателя
    invoice = await session.get(Invoice, callback_data.invoice_id)
    invoice.consignee_id = callback_data.company_id
    await session.commit()

    # --- ИЗМЕНЕНИЕ ЗДЕСЬ ---
    # Достаем контракты ТОЛЬКО для выбранного ранее Отправителя (invoice.company_id)
    contracts_stmt = select(Contract).where(Contract.company_id == invoice.company_id)
    contracts = (await session.scalars(contracts_stmt)).all()

    if not contracts:
        await call.message.edit_text("⚠️ Для данного Отправителя нет контрактов! Добавьте их через /add_contract")
        return

    builder = []
    for cont in contracts:
        builder.append([
            InlineKeyboardButton(
                text=f"📜 №{cont.contract_number} от {cont.contract_date}",
                callback_data=ContractCb(invoice_id=callback_data.invoice_id, contract_id=cont.id).pack()
            )
        ])
    kb = InlineKeyboardMarkup(inline_keyboard=builder)

    await call.message.edit_text(
        "👇 **Шаг 3/3: Выберите Контракт**",
        reply_markup=kb
    )
    await call.answer()

# 3. Выбрали Контракт -> ФИНАЛ
# 3. Выбрали Контракт -> ФИНАЛ И ВЫВОД КНОПОК
@declarant_router.callback_query(ContractCb.filter())
async def select_contract(call: CallbackQuery, callback_data: ContractCb, session):
    invoice = await session.get(Invoice, callback_data.invoice_id)
    invoice.contract_id = callback_data.contract_id
    invoice.status = 'completed'
    await session.commit()

    # Создаем кнопки для скачивания
    builder = [
        [
            InlineKeyboardButton(text="🚚 CMR (Word)", callback_data=DocCb(doc_type="cmr", invoice_id=invoice.id).pack()),
            InlineKeyboardButton(text="📄 Спецификация", callback_data=DocCb(doc_type="spec", invoice_id=invoice.id).pack())
        ],
        [
            InlineKeyboardButton(text="📑 Заявка СТ-1", callback_data=DocCb(doc_type="st1", invoice_id=invoice.id).pack()),
            InlineKeyboardButton(text="📦 Декларация (3qadam)", callback_data=DocCb(doc_type="3qadam", invoice_id=invoice.id).pack())
        ]
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=builder)

    await call.message.edit_text(
        f"🎉 **ИНВОЙС № {invoice.invoice_number} ПОЛНОСТЬЮ ГОТОВ!**\n\n"
        f"Все данные успешно связаны. Выберите нужный документ для генерации:",
        reply_markup=kb
    )
    await call.answer()


# --- ОТКРЫТИЕ ВЫБРАННОГО ИНВОЙСА ИЗ СПИСКА ---
@declarant_router.callback_query(InvoiceListCb.filter())
async def open_invoice_docs(call: CallbackQuery, callback_data: InvoiceListCb, session):
    invoice = await session.get(Invoice, callback_data.invoice_id)

    if not invoice:
        await call.answer("❌ Инвойс не найден или был удален.", show_alert=True)
        return

    # Генерируем те самые 4 кнопки для скачивания
    builder = [
        [
            InlineKeyboardButton(text="🚚 CMR (Word)",
                                 callback_data=DocCb(doc_type="cmr", invoice_id=invoice.id).pack()),
            InlineKeyboardButton(text="📄 Спецификация",
                                 callback_data=DocCb(doc_type="spec", invoice_id=invoice.id).pack())
        ],
        [
            InlineKeyboardButton(text="📑 Заявка СТ-1",
                                 callback_data=DocCb(doc_type="st1", invoice_id=invoice.id).pack()),
            InlineKeyboardButton(text="📦 Декларация (3qadam)",
                                 callback_data=DocCb(doc_type="3qadam", invoice_id=invoice.id).pack())
        ]
    ]
    kb = InlineKeyboardMarkup(inline_keyboard=builder)

    date_str = invoice.invoice_date.strftime('%d.%m.%Y') if invoice.invoice_date else 'Нет данных'

    # Меняем сообщение на карточку инвойса
    await call.message.edit_text(
        f"📂 **Инвойс № {invoice.invoice_number}**\n"
        f"📅 Дата: {date_str}\n"
        f"💰 Общая сумма: {invoice.grand_total} USD\n\n"
        f"👇 Выберите документ для скачивания:",
        reply_markup=kb
    )
    await call.answer()


# 4. Обработка запроса на генерацию документа
@declarant_router.callback_query(DocCb.filter())
async def generate_document_handler(call: CallbackQuery, callback_data: DocCb, session, bot: Bot):
    await call.answer()
    wait_msg = await call.message.answer("⏳ Генерирую документ, подождите...")

    invoice_id = callback_data.invoice_id
    doc_type = callback_data.doc_type

    try:
        # === ВАЖНО: Подгружаем инвойс со ВСЕМИ связанными данными (компания, контракт, товары) ===
        stmt = (
            select(Invoice)
            .where(Invoice.id == invoice_id)
            .options(
                selectinload(Invoice.company),
                selectinload(Invoice.consignee),
                selectinload(Invoice.contract),
                selectinload(Invoice.items)
            )
        )
        invoice = await session.scalar(stmt)

        if not invoice:
            await call.message.answer("❌ Инвойс не найден.")
            await wait_msg.delete()
            return

        # Берем список товаров отдельно для передачи в функцию
        items = invoice.items

        # === Запуск генерации в отдельном потоке, чтобы бот не зависал ===
        if doc_type == "st1":
            filepath = await asyncio.to_thread(generate_st1_file, invoice, items)
            caption_text = f"✅ Ваша заявка СТ-1 для инвойса №{invoice.invoice_number} готова!"
        elif doc_type == "3qadam":
            filepath = await asyncio.to_thread(generate_3qadam_file, invoice, items)
            caption_text = f"✅ Данные декларации (3qadam) для инвойса №{invoice.invoice_number} готовы!"
        elif doc_type == "cmr":
            filepath = await asyncio.to_thread(generate_cmr_file, invoice, items)
            caption_text = f"✅ CMR (Word) для инвойса №{invoice.invoice_number} готова!"
        elif doc_type == "spec":
            filepath = await asyncio.to_thread(generate_spec_file, invoice, items)
            caption_text = f"✅ Спецификация для инвойса №{invoice.invoice_number} готова!"
        else:
            raise ValueError("Неизвестный тип документа")

        # Отправляем сгенерированный файл пользователю в Telegram
        if os.path.exists(filepath):
            document = FSInputFile(filepath)
            await bot.send_document(
                chat_id=call.from_user.id,
                document=document,
                caption=caption_text
            )
            # Удаляем готовый файл после отправки, чтобы не забивать диск сервера
            os.remove(filepath)
        else:
            await call.message.answer("❌ Файл не был создан физически.")

    except Exception as e:
        await call.message.answer(f"❌ Ошибка при генерации документа: {e}")
    finally:
        await wait_msg.delete()


@declarant_router.callback_query(TemplateCb.filter(F.action == "download"))
async def download_template_handler(call: CallbackQuery, bot: Bot):
    await call.answer()  # Сообщаем Telegram, что нажатие принято

    temp_dir = "temp_files"
    os.makedirs(temp_dir, exist_ok=True)
    template_path = os.path.join(temp_dir, "Items_Template.xlsx")

    # Создаем чистый Excel-файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    # Прописываем заголовки строго по нашим колонкам (от 0 до 10)
    headers = [
        "Наименование товара",  # row[0]
        "ТН ВЭД",  # row[1]
        "Кол-во",  # row[2]
        "Ед. изм.",  # row[3]
        "Цена",  # row[4]
        "Кол-во мест",  # row[5]
        "Тип упаковки",  # row[6]
        "Код упаковки (CT, BX)",  # row[7]
        "Штук в упаковке",  # row[8]
        "Вес Нетто",  # row[9]
        "Вес Брутто"  # row[10]
    ]

    # Записываем шапку в первую строку и делаем её жирной (по желанию)
    from openpyxl.styles import Font
    bold_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

    wb.save(template_path)

    # Отправляем пользователю
    document = FSInputFile(template_path)
    await bot.send_document(
        chat_id=call.from_user.id,
        document=document,
        caption="📥 Чистый шаблон для загрузки товаров.\nЗаполните его начиная со 2-й строки, сохраните и отправьте мне в ответ."
    )

    # Удаляем временный файл с сервера
    if os.path.exists(template_path):
        os.remove(template_path)