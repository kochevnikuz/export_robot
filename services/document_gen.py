# services/document_gen.py
import os
import openpyxl
from docxtpl import DocxTemplate
from datetime import date
from datetime import datetime

# Папки для шаблонов и готовых файлов
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_DIR = os.path.join(BASE_DIR, "doc_templates")
OUTPUT_DIR = os.path.join(BASE_DIR, "temp_files")

os.makedirs(TEMPLATE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Словарь для расшифровки кодов стран в CMR
COUNTRY_MAP = {
    "860": "Республика Узбекистан",
    "643": "Российская Федерация",
    "031": "Азербайджан",
    "792": "Турция",
    "156": "Китай",
    "398": "Казахстан",
    "417": "Кыргызстан",
    "762": "Таджикистан",
    "795": "Туркменистан",
    "112": "Беларусь",
    "276": "Германия",
    "380": "Италия",
    "440": "Литва",
    "428": "Латвия",
    "616": "Польша"
}

def generate_st1_file(invoice, items) -> str:
    """Генерация заявки на СТ-1 (Excel)"""
    template_path = os.path.join(TEMPLATE_DIR, 'st1_template.xlsx')
    output_path = os.path.join(OUTPUT_DIR, f"ST1_App_{invoice.invoice_number}.xlsx")

    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    start_row = 2
    formatted_date = invoice.invoice_date.strftime("%d.%m.%Y") if invoice.invoice_date else ""

    for i, item in enumerate(items):
        row = start_row + i

        ws.cell(row=row, column=1, value=item.item_name)
        ws.cell(row=row, column=2, value=item.hs_code)
        ws.cell(row=row, column=3, value=item.quantity)
        ws.cell(row=row, column=4, value=item.unit)
        ws.cell(row=row, column=5, value=item.gross_weight)
        ws.cell(row=row, column=6, value=item.net_weight)
        ws.cell(row=row, column=7, value=166)  # Фиксированный код веса
        ws.cell(row=row, column=8, value=item.package_quantity)
        ws.cell(row=row, column=10, value=item.package_type)
        ws.cell(row=row, column=11, value="Инвойс")  # Фиксированный тип
        ws.cell(row=row, column=12, value=invoice.invoice_number)
        ws.cell(row=row, column=13, value=formatted_date)

    wb.save(output_path)
    return output_path


def generate_3qadam_file(invoice, items) -> str:
    """Генерация данных для декларации 3qadam (Excel)"""
    template_path = os.path.join(TEMPLATE_DIR, '3qadam.xlsx')
    output_path = os.path.join(OUTPUT_DIR, f"Declaration_3qadam_{invoice.invoice_number}.xlsx")

    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    start_row = 4

    # Данные компании (Отправителя) из связи инвойса
    comp_country = invoice.company.country_code if invoice.company else "860"
    comp_tax = invoice.company.tax_id if invoice.company else ""
    comp_area = invoice.company.area_code if invoice.company else ""

    for i, item in enumerate(items):
        row = start_row + i
        num = i + 1

        ws.cell(row=row, column=1, value=num)  # item_number
        # column 2 - пропуск
        ws.cell(row=row, column=3, value=num)  # item_number дубль
        ws.cell(row=row, column=4, value=item.hs_code)
        ws.cell(row=row, column=5, value=item.item_name)
        ws.cell(row=row, column=6, value="БЕЗ МАРКИ")  # brand константа
        ws.cell(row=row, column=7, value=comp_country)
        ws.cell(row=row, column=8, value=comp_tax)
        ws.cell(row=row, column=9, value=comp_area)
        # column 10, 11 - пропуск
        ws.cell(row=row, column=12, value=item.package_code)
        ws.cell(row=row, column=13, value=item.package_quantity)
        # column 14 - пропуск
        ws.cell(row=row, column=15, value=item.gross_weight)
        ws.cell(row=row, column=16, value=item.net_weight)
        # column 17 - пропуск
        ws.cell(row=row, column=18, value=item.total)

    wb.save(output_path)
    return output_path


def generate_cmr_file(invoice, items) -> str:
    """Генерация CMR в Word с группировкой товаров для Telegram-бота"""
    template_path = os.path.join(TEMPLATE_DIR, 'cmr_template.docx')
    output_path = os.path.join(OUTPUT_DIR, f"CMR_{invoice.invoice_number}.docx")

    # Если шаблона еще нет, создаем пустой файл, чтобы код не упал
    if not os.path.exists(template_path):
        from docx import Document
        doc = Document()
        doc.add_paragraph("Шаблон CMR не найден. Пожалуйста, создайте файл cmr_template.docx")
        doc.save(output_path)
        return output_path

    doc = DocxTemplate(template_path)

    # 1. Логика группировки товаров
    grouped_items = {}
    total_places = 0
    total_gross = 0.0

    for item in items:
        # Получаем HS код (или пустую строку, если его нет)
        hs_key = str(item.hs_code).strip() if item.hs_code else "NO_CODE"

        # Получаем первое слово названия
        first_word_name = str(item.item_name).split()[0] if item.item_name else "Товар"

        # Если такого кода еще нет в словаре, создаем запись
        if hs_key not in grouped_items:
            grouped_items[hs_key] = {
                'name': first_word_name,
                'hs': item.hs_code,
                'pack_type': item.package_type,
                'places': 0,
                'gross': 0.0
            }

        # Суммируем данные
        places = int(item.package_quantity) if item.package_quantity else 0
        gross = float(item.gross_weight) if item.gross_weight else 0.0

        grouped_items[hs_key]['places'] += places
        grouped_items[hs_key]['gross'] += gross
        total_places += places
        total_gross += gross

    # Превращаем словарь обратно в список для шаблона Word
    final_items_list = []
    for key, data in grouped_items.items():
        final_items_list.append({
            'name': data['name'],
            'hs': data['hs'],
            'pack_type': data['pack_type'],
            'places': data['places'],
            'gross': round(data['gross'], 2)
        })

        # ... (код выше остается) ...

        # Расшифровываем страны через словарь
        sender_code = str(invoice.company.country_code) if invoice.company else ""
        sender_country_name = COUNTRY_MAP.get(sender_code, sender_code)  # Если кода нет в словаре, оставит как есть

        consignee_code = str(invoice.consignee.country_code) if invoice.consignee else ""
        consignee_country_name = COUNTRY_MAP.get(consignee_code, consignee_code)

        incoterms = invoice.contract.incoterms if invoice.contract else ""
        incoterms_place = invoice.contract.incoterms_place if invoice.contract else ""

        # Достаем cmr_13 из получателя
        cmr_13 = invoice.consignee.cmr_13 if invoice.consignee and invoice.consignee.cmr_13 else ""

        # 2. Собираем контекст
        context = {
            'sender_name': invoice.company.name if invoice.company else "",
            'sender_country': sender_country_name,  # <--- СТРОКА ВМЕСТО ЦИФР
            'sender_address': invoice.company.legal_address if invoice.company else "",

            'consignee_name': invoice.consignee.name if invoice.consignee else "",
            'consignee_address': invoice.consignee.legal_address if invoice.consignee else "",
            'consignee_country': consignee_country_name,  # <--- СТРОКА ВМЕСТО ЦИФР

            'delivery_place': invoice.final_destination or "",

            'transport_number': invoice.transport_number or "",
            'transport_driver': invoice.transport_driver or "",
            'transport_tir_carnet': invoice.transport_tir_carnet or "Нет",

            'invoice_number': invoice.invoice_number,
            'invoice_date': invoice.invoice_date.strftime("%d.%m.%Y") if invoice.invoice_date else "",

            'cmr_13': cmr_13,  # <--- НАСТОЯЩИЕ ДАННЫЕ

            'total_places': total_places,
            'total_gross': round(total_gross, 2),

            'incoterms': incoterms,
            'incoterms_place': incoterms_place,
            'date': date.today().strftime("%d.%m.%Y"),
            'cmr_number': date.today().strftime("%d%m"),

            'items': final_items_list
        }
        # ... (код ниже остается) ...

    # 3. Рендеринг и сохранение файла на диск
    doc.render(context)
    doc.save(output_path)

    # Возвращаем путь к готовому файлу, чтобы бот мог его отправить
    return output_path


def generate_spec_file(invoice, items) -> str:
    """Генерация Спецификации в Excel для Telegram-бота"""
    template_path = os.path.join(TEMPLATE_DIR, 'spec_template.xlsx')
    output_path = os.path.join(OUTPUT_DIR, f"Specification_{invoice.invoice_number}.xlsx")

    # Открываем твой шаблон. Если его случайно нет - создаем чистый
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active

    # 1. Заполняем шапку (Используем данные из нашей БД SQLAlchemy)
    # Если в шаблоне есть эти ячейки, раскомментируй код ниже:

    # contract_number = invoice.contract.contract_number if invoice.contract else ""
    # ws['B2'] = f"Спецификация к Контракту № {contract_number}"

    # invoice_date = invoice.invoice_date.strftime("%d.%m.%Y") if invoice.invoice_date else ""
    # ws['B3'] = f"Инвойс № {invoice.invoice_number} от {invoice_date}"

    # ws['C5'] = invoice.consignee.name if invoice.consignee else ""

    # 2. Заполняем товары
    start_row = 6
    total_amount = 0.0
    total_quantity = 0.0

    # Вместо invoice.items.all() используем список items, который мы передали из хэндлера
    for i, item in enumerate(items):
        row = start_row + i

        # Строго по твоим колонкам: 3(C), 4(D), 5(E), 6(F), 7(G), 9(I)
        ws.cell(row=row, column=3, value=item.item_number)
        ws.cell(row=row, column=4, value=item.hs_code)
        ws.cell(row=row, column=5, value=item.item_name)
        ws.cell(row=row, column=6, value=item.unit)
        ws.cell(row=row, column=7, value=float(item.quantity) if item.quantity else 0.0)
        ws.cell(row=row, column=9, value=float(item.total) if item.total else 0.0)

        total_amount += float(item.total) if item.total else 0.0
        total_quantity += float(item.quantity) if item.quantity else 0.0

    # Итого
    totals_row = start_row + len(items) + 1
    ws.cell(row=totals_row, column=9, value=total_amount)
    ws.cell(row=totals_row, column=7, value=total_quantity)

    # 3. Сохраняем файл на диск и возвращаем путь
    wb.save(output_path)
    return output_path